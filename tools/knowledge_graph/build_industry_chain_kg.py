#!/usr/bin/env python3
"""Build knowledge graph JSON from 产业链图谱.xlsx.

Graph hierarchy:
- Branch (sheet name)
- Sub-chain
- Component
- Product (supports product->product child relation by level code)
- Enterprise

Output:
- tools/knowledge_graph/industry_chain_kg.json
"""

from __future__ import annotations

import argparse
import json
from datetime import datetime
from pathlib import Path
from typing import Any

from openpyxl import load_workbook


DEFAULT_NODE_COLOR = "#90A4AE"
TYPE_COLOR_MAP = {
    "branch": "#546E7A",
    "subchain": "#455A64",
    "component": "#607D8B",
    "product": "#78909C",
    "enterprise": "#8D6E63",
}

LOCALIZATION_COLOR_MAP = {
    "领跑": "#2E7D32",
    "并跑": "#1976D2",
    "跟跑": "#FBC02D",
    "弱项": "#FB8C00",
    "卡脖子": "#D32F2F",
}

ENTERPRISE_TYPE_COLOR_MAP = {
    "国企": "#2E7D32",
    "央企": "#1976D2",
    "民企": "#FBC02D",
    "外企": "#FB8C00",
}


def _norm(value: Any) -> str:
    if value is None:
        return ""
    return str(value).strip()


def _find_key_by_contains(mapping: dict[str, str], raw: str) -> str:
    text = _norm(raw)
    for key, color in mapping.items():
        if key in text:
            return color
    return ""


class GraphBuilder:
    def __init__(self) -> None:
        self.nodes: dict[str, dict[str, Any]] = {}
        self.edges: dict[tuple[str, str, str], dict[str, Any]] = {}

    def add_node(
        self,
        node_id: str,
        label: str,
        node_type: str,
        color: str = "",
        info: dict[str, Any] | None = None,
    ) -> None:
        info = info or {}
        if node_id not in self.nodes:
            self.nodes[node_id] = {
                "id": node_id,
                "label": label,
                "type": node_type,
                "color": color or TYPE_COLOR_MAP.get(node_type, DEFAULT_NODE_COLOR),
                "info": {k: v for k, v in info.items() if _norm(v)},
                "clickable": True,
            }
            return

        # Merge node info if node already exists.
        node = self.nodes[node_id]
        if color:
            node["color"] = color
        for k, v in info.items():
            if _norm(v):
                node["info"][k] = v

    def add_edge(self, source: str, target: str, relation: str) -> None:
        key = (source, target, relation)
        if key in self.edges:
            return
        self.edges[key] = {
            "source": source,
            "target": target,
            "relation": relation,
        }


def _detect_header_row(rows: list[list[str]]) -> tuple[int, dict[str, int]]:
    required_headers = {"子链", "组件", "产品", "产品制造企业"}

    for idx, row in enumerate(rows):
        header_to_idx: dict[str, int] = {}
        for col_idx, cell in enumerate(row):
            text = _norm(cell)
            if text:
                header_to_idx[text] = col_idx
        if required_headers.issubset(set(header_to_idx)):
            return idx, header_to_idx

    raise ValueError("未找到包含关键列(子链/组件/产品/产品制造企业)的表头行")


def build_graph(input_file: Path) -> dict[str, Any]:
    wb = load_workbook(filename=input_file, data_only=True)
    graph = GraphBuilder()

    all_sheet_names = wb.sheetnames

    for sheet_name in all_sheet_names:
        ws = wb[sheet_name]
        raw_rows: list[list[str]] = [
            [_norm(cell) for cell in row]
            for row in ws.iter_rows(values_only=True)
        ]

        if not raw_rows:
            continue

        try:
            header_row_idx, header_map = _detect_header_row(raw_rows)
        except ValueError:
            # Sheet without full data table: still keep branch node.
            branch_id = f"branch|{sheet_name}"
            graph.add_node(
                node_id=branch_id,
                label=sheet_name,
                node_type="branch",
                info={"branch_name": sheet_name},
            )
            continue

        branch_id = f"branch|{sheet_name}"
        graph.add_node(
            node_id=branch_id,
            label=sheet_name,
            node_type="branch",
            info={"branch_name": sheet_name},
        )

        hierarchy_fill_cols = [
            "子链",
            "组件",
            "产品",
            "所属层级",
            "国产化水平",
            "产品整体国产化水平",
            "产品制造企业",
            "产品制造企业类别",
        ]
        detail_cols = [
            "工艺",
            "工器具",
            "软件",
            "卡堵点所处产业链环节",
            "卡堵点",
            "问题具体描述",
            "解决思路",
        ]

        last_seen_hierarchy: dict[str, str] = {k: "" for k in hierarchy_fill_cols}
        last_seen_detail: dict[str, str] = {k: "" for k in detail_cols}
        product_by_level: dict[str, str] = {}
        last_product_group: tuple[str, str, str, str] | None = None

        for row in raw_rows[header_row_idx + 1 :]:
            if not any(_norm(x) for x in row):
                continue

            def value_of(col_name: str) -> str:
                idx = header_map.get(col_name)
                if idx is None or idx >= len(row):
                    return ""
                return _norm(row[idx])

            row_values: dict[str, str] = {}
            for col in hierarchy_fill_cols:
                value = value_of(col)
                if value:
                    last_seen_hierarchy[col] = value
                    row_values[col] = value
                else:
                    row_values[col] = last_seen_hierarchy[col]

            product_group = (
                row_values["子链"],
                row_values["组件"],
                row_values["产品"],
                row_values["所属层级"],
            )

            # Detail fields are inherited only inside the same product group.
            for col in detail_cols:
                value = value_of(col)
                if value:
                    last_seen_detail[col] = value
                    row_values[col] = value
                else:
                    if last_product_group is not None and product_group == last_product_group:
                        row_values[col] = last_seen_detail[col]
                    else:
                        row_values[col] = ""

            subchain = row_values["子链"]
            component = row_values["组件"]
            product = row_values["产品"]
            level_code = row_values["所属层级"]
            enterprise = row_values["产品制造企业"]
            enterprise_type = row_values["产品制造企业类别"]

            if not (subchain or component or product or enterprise):
                continue

            subchain_id = f"subchain|{sheet_name}|{subchain}" if subchain else ""
            component_id = (
                f"component|{sheet_name}|{subchain}|{component}" if component else ""
            )
            product_id = (
                f"product|{sheet_name}|{subchain}|{component}|{level_code}|{product}"
                if product
                else ""
            )
            enterprise_id = f"enterprise|{enterprise}" if enterprise else ""

            subchain_color = _find_key_by_contains(LOCALIZATION_COLOR_MAP, subchain)
            component_color = _find_key_by_contains(
                LOCALIZATION_COLOR_MAP,
                row_values["国产化水平"] or row_values["产品整体国产化水平"],
            )
            product_color = _find_key_by_contains(
                LOCALIZATION_COLOR_MAP,
                row_values["产品整体国产化水平"] or row_values["国产化水平"],
            )
            enterprise_color = _find_key_by_contains(
                ENTERPRISE_TYPE_COLOR_MAP,
                enterprise_type,
            )

            if subchain_id:
                graph.add_node(
                    node_id=subchain_id,
                    label=subchain,
                    node_type="subchain",
                    color=subchain_color,
                    info={
                        "branch_name": sheet_name,
                        "subchain_name": subchain,
                    },
                )
                graph.add_edge(branch_id, subchain_id, "HAS_SUBCHAIN")

            if component_id:
                graph.add_node(
                    node_id=component_id,
                    label=component,
                    node_type="component",
                    color=component_color,
                    info={
                        "branch_name": sheet_name,
                        "subchain_name": subchain,
                        "component_name": component,
                        "localization_level": row_values["国产化水平"],
                    },
                )
                if subchain_id:
                    graph.add_edge(subchain_id, component_id, "HAS_COMPONENT")

            if product_id:
                graph.add_node(
                    node_id=product_id,
                    label=product,
                    node_type="product",
                    color=product_color,
                    info={
                        "branch_name": sheet_name,
                        "subchain_name": subchain,
                        "component_name": component,
                        "product_name": product,
                        "level_code": level_code,
                        "localization_level": row_values["国产化水平"],
                        "overall_localization_level": row_values["产品整体国产化水平"],
                        "process": row_values["工艺"],
                        "tooling": row_values["工器具"],
                        "software": row_values["软件"],
                        "bottleneck_stage": row_values["卡堵点所处产业链环节"],
                        "bottleneck": row_values["卡堵点"],
                        "problem_description": row_values["问题具体描述"],
                        "solution": row_values["解决思路"],
                    },
                )

                parent_added = False
                if level_code and "." in level_code:
                    parent_level = level_code.rsplit(".", 1)[0]
                    parent_product_id = product_by_level.get(parent_level)
                    if parent_product_id and parent_product_id != product_id:
                        graph.add_edge(parent_product_id, product_id, "HAS_PRODUCT")
                        parent_added = True

                if not parent_added and component_id:
                    graph.add_edge(component_id, product_id, "HAS_PRODUCT")

                if level_code:
                    product_by_level[level_code] = product_id

            if enterprise_id:
                graph.add_node(
                    node_id=enterprise_id,
                    label=enterprise,
                    node_type="enterprise",
                    color=enterprise_color,
                    info={
                        "enterprise_name": enterprise,
                        "enterprise_type": enterprise_type,
                    },
                )
                if product_id:
                    graph.add_edge(product_id, enterprise_id, "MANUFACTURED_BY")

            last_product_group = product_group

    nodes = list(graph.nodes.values())
    edges = list(graph.edges.values())

    return {
        "meta": {
            "source_file": str(input_file),
            "generated_at": datetime.now().isoformat(timespec="seconds"),
            "sheet_count": len(all_sheet_names),
            "node_count": len(nodes),
            "edge_count": len(edges),
            "node_color_rules": {
                "localization_level": LOCALIZATION_COLOR_MAP,
                "enterprise_type": ENTERPRISE_TYPE_COLOR_MAP,
                "type_default": TYPE_COLOR_MAP,
            },
        },
        "nodes": nodes,
        "edges": edges,
    }


def write_graph_html(graph_json: dict[str, Any], output_html: Path) -> None:
        nodes_json = json.dumps(graph_json["nodes"], ensure_ascii=False)
        edges_json = json.dumps(graph_json["edges"], ensure_ascii=False)
        meta_json = json.dumps(graph_json["meta"], ensure_ascii=False)

        html = f"""<!doctype html>
<html lang="zh-CN">
<head>
    <meta charset="utf-8" />
    <meta name="viewport" content="width=device-width, initial-scale=1" />
    <title>产业链知识图谱</title>
    <script src="https://unpkg.com/vis-network@9.1.2/dist/vis-network.min.js"></script>
    <style>
        :root {{
            --bg: #f3f6f7;
            --panel: #ffffff;
            --border: #dce3e6;
            --text: #1f2a30;
            --muted: #5a6a73;
            --accent: #0e7490;
        }}
        * {{ box-sizing: border-box; }}
        body {{
            margin: 0;
            font-family: "PingFang SC", "Microsoft YaHei", "Noto Sans SC", sans-serif;
            color: var(--text);
            background: radial-gradient(1200px 500px at 10% 0%, #e8f4f6 0%, var(--bg) 65%);
        }}
        .layout {{
            display: grid;
            grid-template-columns: 300px 1fr 360px;
            gap: 12px;
            padding: 12px;
            height: 100vh;
        }}
        .card {{
            background: var(--panel);
            border: 1px solid var(--border);
            border-radius: 12px;
            overflow: hidden;
            box-shadow: 0 4px 14px rgba(14, 116, 144, 0.06);
        }}
        .card h3 {{
            margin: 0;
            padding: 12px 14px;
            font-size: 15px;
            border-bottom: 1px solid var(--border);
            background: #f9fbfc;
        }}
        .card .body {{
            padding: 12px 14px;
            overflow: auto;
            height: calc(100% - 46px);
        }}
        #network {{
            width: 100%;
            height: 100%;
            min-height: 520px;
        }}
        .legend-row {{ display: flex; align-items: center; margin-bottom: 8px; font-size: 13px; color: var(--muted); }}
        .dot {{ width: 12px; height: 12px; border-radius: 999px; margin-right: 8px; border: 1px solid #c9d3d8; }}
        .meta {{ font-size: 13px; color: var(--muted); line-height: 1.7; }}
        .search-wrap {{ display: flex; gap: 8px; margin-bottom: 12px; }}
        input, button {{ font: inherit; }}
        input[type="text"] {{
            width: 100%;
            padding: 8px 10px;
            border: 1px solid var(--border);
            border-radius: 8px;
            outline: none;
            background: #fff;
        }}
        button {{
            border: 1px solid var(--border);
            background: #fff;
            color: var(--text);
            border-radius: 8px;
            padding: 8px 10px;
            cursor: pointer;
        }}
        button.primary {{
            background: var(--accent);
            border-color: var(--accent);
            color: #fff;
        }}
        .info-title {{ font-weight: 600; margin-bottom: 8px; }}
        .kv {{ font-size: 13px; margin-bottom: 8px; }}
        .kv .k {{ color: var(--muted); }}
        .empty {{ color: var(--muted); font-size: 13px; }}
        @media (max-width: 1180px) {{
            .layout {{ grid-template-columns: 1fr; height: auto; }}
            .card {{ min-height: 280px; }}
            #network {{ min-height: 520px; }}
        }}
    </style>
</head>
<body>
    <div class="layout">
        <div class="card">
            <h3>图谱说明</h3>
            <div class="body">
                <div class="meta" id="meta"></div>
                <hr style="border:none;border-top:1px solid var(--border);margin:12px 0;" />
                <div style="font-size:13px;font-weight:600;margin-bottom:8px;">国产化水平颜色</div>
                <div class="legend-row"><span class="dot" style="background:#2E7D32"></span>领跑</div>
                <div class="legend-row"><span class="dot" style="background:#1976D2"></span>并跑</div>
                <div class="legend-row"><span class="dot" style="background:#FBC02D"></span>跟跑</div>
                <div class="legend-row"><span class="dot" style="background:#FB8C00"></span>弱项</div>
                <div class="legend-row"><span class="dot" style="background:#D32F2F"></span>卡脖子</div>
                <hr style="border:none;border-top:1px solid var(--border);margin:12px 0;" />
                <div style="font-size:13px;font-weight:600;margin-bottom:8px;">企业类型颜色</div>
                <div class="legend-row"><span class="dot" style="background:#2E7D32"></span>国企</div>
                <div class="legend-row"><span class="dot" style="background:#1976D2"></span>央企</div>
                <div class="legend-row"><span class="dot" style="background:#FBC02D"></span>民企</div>
                <div class="legend-row"><span class="dot" style="background:#FB8C00"></span>外企</div>
            </div>
        </div>

        <div class="card">
            <h3>产业链知识图谱</h3>
            <div class="body" style="padding:10px;">
                <div class="search-wrap">
                    <input id="searchInput" type="text" placeholder="输入节点名称，如 IGBT / 中车 / 柔直换流阀" />
                    <button class="primary" id="searchBtn">定位</button>
                    <button id="fitBtn">全图</button>
                </div>
                <div id="network"></div>
            </div>
        </div>

        <div class="card">
            <h3>节点详情</h3>
            <div class="body" id="detailPanel">
                <div class="empty">点击任意节点查看详细信息。</div>
            </div>
        </div>
    </div>

    <script>
        const graphMeta = {meta_json};
        const rawNodes = {nodes_json};
        const rawEdges = {edges_json};

        const nodes = new vis.DataSet(rawNodes.map(n => {{
            const title = `${{n.type}}: ${{n.label}}`;
            return {{
                id: n.id,
                label: n.label,
                color: n.color,
                title,
                shape: "dot",
                size: n.type === "branch" ? 18 : n.type === "subchain" ? 14 : 10,
                font: {{ size: 13, color: "#1f2a30" }},
            }};
        }}));

        const edges = new vis.DataSet(rawEdges.map(e => ({{
            from: e.source,
            to: e.target,
            arrows: "to",
            color: {{ color: "#9fb0b9", opacity: 0.8 }},
            width: 1.2,
            smooth: {{ type: "cubicBezier", roundness: 0.2 }},
        }})));

        const container = document.getElementById("network");
        const network = new vis.Network(
            container,
            {{ nodes, edges }},
            {{
                physics: {{
                    stabilization: false,
                    barnesHut: {{ gravitationalConstant: -4000, springLength: 140, springConstant: 0.03 }}
                }},
                interaction: {{ hover: true, multiselect: false }},
            }}
        );

        const detailPanel = document.getElementById("detailPanel");
        const nodeById = new Map(rawNodes.map(n => [n.id, n]));

        document.getElementById("meta").innerHTML = [
            `来源文件: ${{graphMeta.source_file}}`,
            `工作表数量: ${{graphMeta.sheet_count}}`,
            `节点数量: ${{graphMeta.node_count}}`,
            `关系数量: ${{graphMeta.edge_count}}`,
            `生成时间: ${{graphMeta.generated_at}}`,
        ].map(x => `<div>${{x}}</div>`).join("");

        network.on("click", (params) => {{
            if (!params.nodes.length) return;
            const nodeId = params.nodes[0];
            const node = nodeById.get(nodeId);
            if (!node) return;

            const rows = Object.entries(node.info || {{}})
                .map(([k, v]) => `<div class=\"kv\"><span class=\"k\">${{k}}：</span>${{String(v)}}</div>`)
                .join("");

            detailPanel.innerHTML = `
                <div class=\"info-title\">${{node.label}}</div>
                <div class=\"kv\"><span class=\"k\">节点类型：</span>${{node.type}}</div>
                <div class=\"kv\"><span class=\"k\">节点ID：</span>${{node.id}}</div>
                ${{rows || '<div class=\"empty\">无扩展信息</div>'}}
            `;
        }});

        document.getElementById("searchBtn").addEventListener("click", () => {{
            const kw = (document.getElementById("searchInput").value || "").trim().toLowerCase();
            if (!kw) return;
            const hit = rawNodes.find(n => (n.label || "").toLowerCase().includes(kw));
            if (!hit) {{
                alert("未找到匹配节点");
                return;
            }}
            network.selectNodes([hit.id]);
            network.focus(hit.id, {{ scale: 1.1, animation: {{ duration: 500 }} }});
            network.emit("click", {{ nodes: [hit.id] }});
        }});

        document.getElementById("fitBtn").addEventListener("click", () => {{
            network.fit({{ animation: true }});
        }});
    </script>
</body>
</html>
"""

        output_html.parent.mkdir(parents=True, exist_ok=True)
        output_html.write_text(html, encoding="utf-8")


def main() -> None:
    parser = argparse.ArgumentParser(description="Build industry chain knowledge graph JSON")
    parser.add_argument(
        "--input",
        default="产业链图谱.xlsx",
        help="Input xlsx file path",
    )
    parser.add_argument(
        "--output",
        default="tools/knowledge_graph/industry_chain_kg.json",
        help="Output json file path",
    )
    parser.add_argument(
        "--output-html",
        default="tools/knowledge_graph/industry_chain_kg.html",
        help="Output html file path for interactive graph",
    )
    args = parser.parse_args()

    script_dir = Path(__file__).resolve().parent
    workspace_root = script_dir.parent.parent

    input_path = Path(args.input)
    output_path = Path(args.output)
    output_html_path = Path(args.output_html)

    # Allow running from either workspace root or tools/knowledge_graph.
    if not input_path.is_absolute() and not input_path.exists():
        candidate = workspace_root / input_path
        if candidate.exists():
            input_path = candidate

    if not output_path.is_absolute():
        output_path = workspace_root / output_path
    if not output_html_path.is_absolute():
        output_html_path = workspace_root / output_html_path

    if not input_path.exists():
        raise FileNotFoundError(f"Input file not found: {input_path}")

    graph_json = build_graph(input_path)

    output_path.parent.mkdir(parents=True, exist_ok=True)
    output_path.write_text(
        json.dumps(graph_json, ensure_ascii=False, indent=2),
        encoding="utf-8",
    )
    write_graph_html(graph_json, output_html_path)

    meta = graph_json["meta"]
    print("Knowledge graph generated:")
    print(f"- output: {output_path}")
    print(f"- output_html: {output_html_path}")
    print(f"- sheets: {meta['sheet_count']}")
    print(f"- nodes: {meta['node_count']}")
    print(f"- edges: {meta['edge_count']}")


if __name__ == "__main__":
    main()
